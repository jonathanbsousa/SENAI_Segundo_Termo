from django.shortcuts import render
from .serializers import *
from .filters import *
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, parser_classes
from rest_framework.generics import ListCreateAPIView, CreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import filters
from openpyxl import load_workbook, Workbook
from datetime import datetime
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse

class Sensores_List_Create(ListCreateAPIView):
    queryset = Sensores.objects.all()
    serializer_class = Sensores_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SensoresFilter

class Sensores_Retrieve_Update_Destroy(RetrieveUpdateDestroyAPIView):
    queryset = Sensores.objects.all()
    serializer_class = Sensores_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SensoresFilter

class Ambientes_List_Create(ListCreateAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = Ambientes_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = AmbientesFilter

class Ambientes_Retrieve_Update_Destroy(RetrieveUpdateDestroyAPIView):
    queryset = Ambientes.objects.all()
    serializer_class = Ambientes_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = AmbientesFilter

class Historico_List_Create(ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = Historico_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = HistoricoFilter

class Historico_Retrieve_Update_Destroy(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = Historico_Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = HistoricoFilter

class SignUpView(CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class UploadXLSXView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sensor = str(row[0]).strip() if row[0] is not None else None
            mac_address = str(row[1]).strip() if row[1] is not None else None
            unidade_med = str(row[2]).strip() if row[2] is not None else None
            latitude = row[3] if row[3] is not None else None
            longitude = row[4] if row[4] is not None else None
            status = str(row[5]).strip().lower() if row[5] is not None else 'ativo'

            # Validação mínima
            if not sensor or not mac_address:
                print(f"[Linha {i+2}] Erro: campos obrigatórios ausentes. Dados: {row}")
                continue

            if status not in ['ativo', 'inativo']:
                status = 'ativo'  # valor padrão em caso de erro

            try:
                latitude = float(latitude)
                longitude = float(longitude)
            except (TypeError, ValueError):
                print(f"[Linha {i+2}] Erro de conversão em latitude/longitude. Dados: {row}")
                continue

            Sensores.objects.create(
                sensor=sensor,
                mac_address=mac_address,
                unidade_med=unidade_med,
                latitude=latitude,
                longitude=longitude,
                status=status
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})

class UploadHistoricoXLSXView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sensor = str(row[0]).strip() if row[0] else None
            ambiente = str(row[1]).strip() if row[1] else None
            valor = row[2]
            timestamp_raw = row[3]

            # Validação mínima
            if not sensor or not ambiente or valor is None or timestamp_raw is None:
                print(f"[Linha {i+2}] Dados incompletos. Ignorando. Linha: {row}")
                continue

            try:
                sensor = Sensores.objects.get(id=int(sensor))
            except Sensores.DoesNotExist:
                print(f"[Linha {i+2}] Sensor '{sensor}' não encontrado.")
                continue

            try:
                ambiente = Ambientes.objects.get(id=int(ambiente))
            except Ambientes.DoesNotExist:
                print(f"[Linha {i+2}] Ambiente '{ambiente}' não encontrado.")
                continue

            try:
                timestamp = (
                    timestamp_raw if isinstance(timestamp_raw, datetime)
                    else datetime.strptime(str(timestamp_raw), "%Y-%m-%d %H:%M:%S")
                )
            except Exception as e:
                print(f"[Linha {i+2}] Erro ao converter timestamp: {timestamp_raw} ({e})")
                continue

            Historico.objects.create(
                sensor=sensor,
                ambiente=ambiente,
                valor=valor,
                timestamp=timestamp
            )

        return Response({'mensagem': 'Histórico importado com sucesso!'})
    
class UploadAmbientesXLSXView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba da planilha

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sig = row[0]
            descricao = str(row[1]).strip() if row[1] else None
            ni = str(row[2]).strip() if row[2] else None
            responsavel = str(row[3]).strip() if row[3] else None

            if sig is None or not descricao or not ni or not responsavel:
                print(f"[Linha {i+2}] Dados incompletos. Ignorando. Linha: {row}")
                continue

            try:
                sig = int(sig)
            except ValueError:
                print(f"[Linha {i+2}] sig inválido: {sig}")
                continue

            Ambientes.objects.create(
                sig=sig,
                descricao=descricao,
                ni=ni,
                responsavel=responsavel
            )

        return Response({'mensagem': 'Ambientes importados com sucesso!'})

class ExportXLSXHistorico(APIView):
    def get(self, request):
        # Aplica os filtros com base na URL
        filtro = HistoricoFilter(request.GET, queryset=Historico.objects.all())
        historico_filtrado = filtro.qs

        # Cria nova planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"

        # Cabeçalho
        ws.append(['Sensor', 'Ambiente', 'Valor', 'Timestamp'])

        # Dados filtrados
        for item in historico_filtrado:
            ws.append([
                item.sensor.sensor,
                item.ambiente.descricao,
                item.valor,
                item.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            ])

        # Gera o XLSX como resposta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=historico_filtrado.xlsx'
        wb.save(response)

        return response
    
class ExportXLSXSensores(APIView):
    def get(self, request):
        sensores = Sensores.objects.all()
        wb = load_workbook('sensores.xlsx')
        ws = wb.active

        # Cabeçalho
        ws.append(['Sensor', 'Mac Address', 'Unidade Med', 'Latitude', 'Longitude', 'Status'])

        # Dados
        for item in sensores:
            ws.append([
                item.sensor,
                item.mac_address,
                item.unidade_med,
                item.latitude,
                item.longitude,
                item.status
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sensores.xlsx'
        wb.save(response)

        return response
    
class ExportXLSXAmbientes(APIView):
    def get(self, request):
        ambientes = Ambientes.objects.all()
        wb = load_workbook('ambientes.xlsx')
        ws = wb.active

        # Cabeçalho
        ws.append(['Sig', 'Descrição', 'NI', 'Responsável'])

        # Dados
        for item in ambientes:
            ws.append([
                item.sig,
                item.descricao,
                item.ni,
                item.responsavel
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ambientes.xlsx'
        wb.save(response)

        return response